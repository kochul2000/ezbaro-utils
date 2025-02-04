import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import Workbook, load_workbook

class ExcelMergeApp:
    VERSION = "1.0.1"

    def __init__(self, master):
        self.master = master
        self.master.title("엑셀 병합 도구 v" + self.VERSION)

        # ───────────────────────────────────────────
        # 1-1) 옵션/컨트롤 패널 영역 (체크박스 + 헤더 행 수)
        # ───────────────────────────────────────────
        control_frame = tk.Frame(self.master, bd=2, relief='groove')
        control_frame.pack(padx=10, pady=5, fill='x')

        # 스타일 / 하이퍼링크 / 댓글 복사 여부 결정
        self.copy_style_var = tk.BooleanVar(value=True)
        self.copy_hyperlink_var = tk.BooleanVar(value=True)
        self.copy_comment_var = tk.BooleanVar(value=True)

        tk.Checkbutton(control_frame, text="스타일 복사", variable=self.copy_style_var)\
          .grid(row=0, column=0, padx=5, pady=5, sticky='w')
        tk.Checkbutton(control_frame, text="하이퍼링크 복사", variable=self.copy_hyperlink_var)\
          .grid(row=0, column=1, padx=5, pady=5, sticky='w')
        tk.Checkbutton(control_frame, text="댓글 복사", variable=self.copy_comment_var)\
          .grid(row=0, column=2, padx=5, pady=5, sticky='w')

        # 헤더 행 수
        tk.Label(control_frame, text="헤더 행 수:", width=10)\
          .grid(row=1, column=0, padx=5, pady=5, sticky='e')
        self.header_count_var = tk.StringVar(value="1")  # 기본값: 1
        header_count_spinbox = tk.Spinbox(control_frame, from_=0, to=100,
                                          textvariable=self.header_count_var, width=5)
        header_count_spinbox.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        # ───────────────────────────────────────────
        # 1-2) 내부 특수 컨트롤
        # ───────────────────────────────────────────
        control_ete_frame = tk.Frame(self.master, bd=2, relief='groove')
        control_ete_frame.pack(padx=10, pady=5, fill='x')
        # 여기서 "파일명 Clear" 버튼 추가
        clear_button = tk.Button(control_ete_frame, text="파일 경로 초기화", command=self.clear_file_paths)
        clear_button.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        # ───────────────────────────────────────────
        # 2) 상단 헤더 프레임(시트명 라벨 + 파일 경로 라벨 + '파일 열 추가' 버튼)
        # ───────────────────────────────────────────
        header_frame = tk.Frame(self.master)
        header_frame.pack(padx=10, pady=5, fill='x')

        tk.Label(header_frame, text="시트명", width=15).grid(row=0, column=0, padx=5, pady=5)

        self.file_col_count = 2
        for col_idx in range(self.file_col_count):
            tk.Label(header_frame, text=f"파일 경로 #{col_idx+1}", width=30)\
              .grid(row=0, column=col_idx+1, padx=5, pady=5)

        self.add_col_button = tk.Button(header_frame, text="파일 열 추가", command=self.add_column)
        self.add_col_button.grid(row=0, column=self.file_col_count + 1, padx=5, pady=5, sticky='w')

        # ───────────────────────────────────────────
        # 3) 메인 프레임(시트명 + 파일 경로 테이블)
        # ───────────────────────────────────────────
        self.main_frame = tk.Frame(self.master)
        self.main_frame.pack(padx=10, pady=5, fill='x')

        # 각 행을 저장할 리스트
        self.sheet_rows = []
        # 초기로 1개 행 추가
        self.add_row()

        # ───────────────────────────────────────────
        # 4) 하단 버튼 프레임 (행 추가, "파일명 Clear", 병합)
        # ───────────────────────────────────────────
        button_frame = tk.Frame(self.master)
        button_frame.pack(padx=10, pady=5, fill='x')

        add_row_button = tk.Button(button_frame, text="행 추가", command=self.add_row)
        add_row_button.pack(side='left', padx=5)

        merge_button = tk.Button(button_frame, text="병합", command=self.merge_excel)
        merge_button.pack(side='right', padx=5)

    def add_row(self):
        """
        (시트명 + 파일경로들) 행을 새로 추가
        """
        row_index = len(self.sheet_rows)

        sheet_entry = tk.Entry(self.main_frame, width=15)
        sheet_entry.grid(row=row_index, column=0, padx=5, pady=5)
        sheet_entry.insert(0, f"시트{row_index+1}")

        file_entries = []
        for col_idx in range(self.file_col_count):
            file_entry = tk.Entry(self.main_frame, width=30)
            file_entry.grid(row=row_index, column=col_idx+1, padx=5, pady=5, sticky='w')

            browse_button = tk.Button(self.main_frame, text="찾아보기",
                                      command=lambda e=file_entry: self.browse_file(e))
            browse_button.grid(row=row_index, column=col_idx+1, padx=5, pady=5, sticky='e')
            file_entries.append(file_entry)

        self.sheet_rows.append({
            'sheet_entry': sheet_entry,
            'files': file_entries
        })

    def add_column(self):
        """
        모든 행(Row)에 대해 새로운 파일 경로 열(Column)을 추가
        """
        frames = list(self.master.children.values())
        # [0] -> control_frame, [1] -> header_frame, [2] -> main_frame, ...
        header_frame = frames[2]

        new_col_idx = self.file_col_count + 1
        tk.Label(header_frame, text=f"파일 경로 #{self.file_col_count+1}", width=30)\
            .grid(row=0, column=new_col_idx, padx=5, pady=5)

        for row_index, row_data in enumerate(self.sheet_rows):
            file_entry = tk.Entry(self.main_frame, width=30)
            file_entry.grid(row=row_index, column=new_col_idx, padx=5, pady=5, sticky='w')

            browse_button = tk.Button(self.main_frame, text="찾아보기",
                                      command=lambda e=file_entry: self.browse_file(e))
            browse_button.grid(row=row_index, column=new_col_idx, padx=5, pady=5, sticky='e')

            row_data['files'].append(file_entry)

        self.file_col_count += 1
        self.add_col_button.grid(row=0, column=self.file_col_count + 1, padx=5, pady=5, sticky='w')

    def browse_file(self, entry_widget):
        """
        파일 대화상자를 통해 파일 경로 선택 후 Entry에 넣기
        """
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=(("Excel 파일", "*.xlsx *.xlsm *.xltx *.xltm"), ("모든 파일", "*.*"))
        )
        if file_path:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, file_path)

    def clear_file_paths(self):
        """
        모든 행의 'files' Entry를 비움
        """
        for row_data in self.sheet_rows:
            for file_entry in row_data['files']:
                file_entry.delete(0, tk.END)

    def merge_excel(self):
        """
        병합 로직
        """
        # 헤더 행 수
        try:
            header_count = int(self.header_count_var.get().strip())
        except ValueError:
            messagebox.showwarning("주의", "헤더 행 수가 잘못되었습니다. 정수를 입력해주세요.")
            return

        if header_count < 0:
            messagebox.showwarning("주의", "헤더 행 수는 0 이상이어야 합니다.")
            return

        # 체크박스 상태
        copy_style = self.copy_style_var.get()
        copy_hyperlink = self.copy_hyperlink_var.get()
        copy_comment = self.copy_comment_var.get()

        # 새 워크북
        wb_new = Workbook()
        default_sheet = wb_new.active
        wb_new.remove(default_sheet)

        has_data = False

        for row_data in self.sheet_rows:
            sheet_name = row_data['sheet_entry'].get().strip()
            file_paths = [f.get().strip() for f in row_data['files'] if f.get().strip()]

            if not file_paths:
                continue

            has_data = True
            # 새 시트
            new_sheet = wb_new.create_sheet(title=sheet_name)

            for idx, path in enumerate(file_paths):
                try:
                    wb_source = load_workbook(path, data_only=False)
                    ws_source = wb_source.active  # 첫 시트
                except Exception as e:
                    messagebox.showerror("에러", f"파일 열기에 실패:\n{path}\n{e}")
                    return

                skip_count = 0 if idx == 0 else header_count

                self.append_sheet_data(
                    ws_source, new_sheet,
                    is_first_file=(idx==0),
                    skip_count=skip_count,
                    copy_style=copy_style,
                    copy_hyperlink=copy_hyperlink,
                    copy_comment=copy_comment
                )

        if not has_data:
            messagebox.showwarning("주의", "병합할 파일이 없습니다.")
            return

        save_path = filedialog.asksaveasfilename(
            title="병합 결과 저장",
            defaultextension=".xlsx",
            filetypes=(("Excel 파일", "*.xlsx"), ("모든 파일", "*.*"))
        )
        if not save_path:
            return

        try:
            wb_new.save(save_path)
            messagebox.showinfo("완료", f"병합이 완료되었습니다!\n{save_path}")
        except Exception as e:
            messagebox.showerror("에러", f"저장 실패:\n{save_path}\n{e}")

    def append_sheet_data(self, ws_source, ws_target,
                          is_first_file=False,
                          skip_count=0,
                          copy_style=True,
                          copy_hyperlink=True,
                          copy_comment=True):
        """
        ws_source → ws_target 로 복사
        """
        max_row = ws_source.max_row
        max_col = ws_source.max_column

        if is_first_file:
            for col_letter, col_dim in ws_source.column_dimensions.items():
                ws_target.column_dimensions[col_letter].width = col_dim.width

        # 빈 시트 체크
        def is_sheet_empty(sheet):
            if sheet.max_row <= 1:
                row1_values = [cell.value for cell in sheet[1]]
                if all(v is None for v in row1_values):
                    return True
            return False

        for row_idx in range(1 + skip_count, max_row + 1):
            if is_sheet_empty(ws_target):
                t_row = 1
            else:
                t_row = ws_target.max_row + 1

            for col_idx in range(1, max_col + 1):
                s_cell = ws_source.cell(row=row_idx, column=col_idx)
                t_cell = ws_target.cell(row=t_row, column=col_idx)

                t_cell.value = s_cell.value

                # 스타일 복사
                if copy_style and s_cell.has_style:
                    t_cell.font = s_cell.font.copy()
                    t_cell.border = s_cell.border.copy()
                    t_cell.fill = s_cell.fill.copy()
                    t_cell.number_format = s_cell.number_format
                    t_cell.protection = s_cell.protection.copy()
                    t_cell.alignment = s_cell.alignment.copy()

                # 하이퍼링크
                if copy_hyperlink and s_cell.hyperlink:
                    t_cell.hyperlink = s_cell.hyperlink

                # 댓글
                if copy_comment and s_cell.comment:
                    t_cell.comment = openpyxl.comments.Comment(
                        text=s_cell.comment.text,
                        author=s_cell.comment.author
                    )

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergeApp(root)
    root.mainloop()
